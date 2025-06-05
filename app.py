import os
import re
import sqlite3
import uuid
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, redirect, url_for, request, session, flash, jsonify
import cv2
import numpy as np
import psutil  # Added psutil import
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import subprocess
import platform
import logging
import json
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import make_response
import io
from datetime import datetime


# Import utility modules
from utils.face_recognition import compare_faces
from utils.wifi_checker import check_wifi_connection, get_current_wifi_info, verify_college_wifi


app = Flask(__name__)
app.secret_key = "your_secret_key_here"
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['ALLOWED_EXTENSIONS'] = {'png', 'jpg', 'jpeg'}

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Ensure upload directory exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Database connection helper
def get_db_connection():
    conn = sqlite3.connect('database/attendance.db')
    conn.row_factory = sqlite3.Row
    return conn

# Initialize the database
def init_db():
    with app.app_context():
        conn = get_db_connection()
        
        # First check if the allowed_access_points table exists
        cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='allowed_access_points'")
        table_exists = cursor.fetchone() is not None
        
        # If the table doesn't exist, create it and add example data
        if not table_exists:
            conn.executescript('''
                -- Add new table to store allowed access points
                CREATE TABLE IF NOT EXISTS allowed_access_points (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    ssid TEXT NOT NULL,
                    bssid TEXT NOT NULL,  -- MAC address of the access point
                    description TEXT,     -- Optional description (e.g., "Library AP", "Classroom Building")
                    UNIQUE(bssid)
                );
                
                -- Insert some example access points (you'll replace these with your actual college APs)
                INSERT OR IGNORE INTO allowed_access_points (ssid, bssid, description) 
                VALUES 
                    ('College_WiFi', '00:11:22:33:44:55', 'Main Building - Floor 1'),
                    ('College_WiFi', '00:11:22:33:44:56', 'Main Building - Floor 2'),
                    ('College_WiFi', '00:11:22:33:44:57', 'Library'),
                    ('College_WiFi', '00:11:22:33:44:58', 'Science Building');
            ''')
            conn.commit()
            print("Created allowed_access_points table")
        
        # Initialize the database from schema if it's a fresh database
        if not conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='users'").fetchone():
            with open('database/schema.sql') as f:
                conn.executescript(f.read())
            print("Initialized database with schema")
        
        # Check if admin exists, if not create one
        cursor = conn.execute('SELECT * FROM users WHERE role = ?', ('admin',))
        if cursor.fetchone() is None:
            # Create default admin (username: admin@college.edu, password: admin123)
            conn.execute(
                'INSERT INTO users (email, password, role) VALUES (?, ?, ?)',
                ('admin@college.edu', generate_password_hash('admin123'), 'admin')
            )
            conn.execute(
                'INSERT INTO settings (wifi_ssid, wifi_password) VALUES (?, ?)',
                ('College_WiFi', 'college_password')
            )
            conn.commit()
            print("Created default admin user")
        
        conn.close()

# Helper function to check if file extension is allowed
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

# Login required decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please login first', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Admin required decorator
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or session.get('role') != 'admin':
            flash('Admin access required', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Helper function to get WiFi MAC address 
def get_wifi_mac():
    try:
        addrs = psutil.net_if_addrs()
        for iface_name, iface_addrs in addrs.items():
            # Look for wifi interfaces
            if any(name in iface_name.lower() for name in ['wi-fi', 'wlan', 'wireless']):
                for addr in iface_addrs:
                    # AF_LINK is the MAC address family type
                    if hasattr(addr, 'family') and (
                        getattr(addr.family, 'name', None) == 'AF_LINK' or 
                        addr.family == psutil.AF_LINK
                    ):
                        return addr.address
        return "Wi-Fi MAC not found"
    except Exception as e:
        app.logger.error(f"Error getting WiFi MAC: {e}")
        return "Error detecting Wi-Fi MAC"

# Routes
@app.route('/')
def index():
    if 'user_id' in session:
        if session['role'] == 'admin':
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for('student_dashboard'))
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        conn = get_db_connection()
        user = conn.execute('SELECT * FROM users WHERE email = ?', (email,)).fetchone()
        conn.close()
        
        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['id']
            session['email'] = user['email']
            session['role'] = user['role']
            
            if user['role'] == 'admin':
                return redirect(url_for('admin_dashboard'))
            else:
                return redirect(url_for('student_dashboard'))
        else:
            flash('Invalid email or password', 'danger')
    
    return render_template('index.html')




@app.route('/admin/attendance-report/export')
@login_required
@admin_required
def export_attendance_report():
    conn = get_db_connection()
    
    # Get all students with their attendance data (same as attendance_report route)
    students_data = conn.execute('''
        SELECT s.id, s.name, s.roll_number, s.email,
               COUNT(a.id) as days_present
        FROM students s
        LEFT JOIN attendance a ON s.id = a.student_id
        GROUP BY s.id, s.name, s.roll_number, s.email
        ORDER BY s.roll_number
    ''').fetchall()
    
    # Calculate total working days
    total_days_result = conn.execute('SELECT COUNT(DISTINCT date) as total_days FROM attendance').fetchone()
    total_working_days = max(total_days_result['total_days'], 1)
    
    if total_working_days == 0:
        total_working_days = 90
    
    conn.close()
    
    # Calculate percentages
    students_with_percentage = []
    for student in students_data:
        percentage = (student['days_present'] / total_working_days) * 100
        students_with_percentage.append({
            'id': student['id'],
            'name': student['name'],
            'roll_number': student['roll_number'],
            'email': student['email'],
            'days_present': student['days_present'],
            'total_days': total_working_days,
            'percentage': round(percentage, 2),
            'is_shortage': percentage < 75
        })
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    summary_font = Font(bold=True, size=12)
    summary_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    shortage_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    good_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title and summary information
    ws.merge_cells('A1:H1')
    ws['A1'] = "ATTENDANCE REPORT"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_alignment
    
    # Add generation date
    ws.merge_cells('A2:H2')
    ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].alignment = center_alignment
    
    # Add summary statistics
    total_students = len(students_with_percentage)
    good_attendance = len([s for s in students_with_percentage if not s['is_shortage']])
    shortage_students = len([s for s in students_with_percentage if s['is_shortage']])
    avg_attendance = sum(s['percentage'] for s in students_with_percentage) / total_students if total_students > 0 else 0
    
    # Summary section
    ws['A4'] = "SUMMARY"
    ws['A4'].font = summary_font
    ws['A4'].fill = summary_fill
    
    ws['A5'] = "Total Working Days:"
    ws['B5'] = total_working_days
    ws['A6'] = "Total Students:"
    ws['B6'] = total_students
    ws['A7'] = "Good Attendance (≥75%):"
    ws['B7'] = good_attendance
    ws['A8'] = "Attendance Shortage (<75%):"
    ws['B8'] = shortage_students
    ws['A9'] = "Average Attendance:"
    ws['B9'] = f"{avg_attendance:.1f}%"
    
    # Style summary section
    for row in range(5, 10):
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = summary_fill
        ws[f'B{row}'].fill = summary_fill
    
    # Add student data table
    start_row = 12
    headers = ['Roll Number', 'Name', 'Email', 'Days Present', 'Total Days', 'Percentage', 'Status']
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Add student data
    for row_idx, student in enumerate(students_with_percentage, start_row + 1):
        ws.cell(row=row_idx, column=1, value=student['roll_number']).border = border
        ws.cell(row=row_idx, column=2, value=student['name']).border = border
        ws.cell(row=row_idx, column=3, value=student['email']).border = border
        ws.cell(row=row_idx, column=4, value=student['days_present']).border = border
        ws.cell(row=row_idx, column=5, value=student['total_days']).border = border
        ws.cell(row=row_idx, column=6, value=f"{student['percentage']}%").border = border
        ws.cell(row=row_idx, column=7, value="Shortage" if student['is_shortage'] else "Good").border = border
        
        # Apply conditional formatting
        if student['is_shortage']:
            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).fill = shortage_fill
        else:
            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).fill = good_fill
        
        # Center align numeric columns
        ws.cell(row=row_idx, column=4).alignment = center_alignment
        ws.cell(row=row_idx, column=5).alignment = center_alignment
        ws.cell(row=row_idx, column=6).alignment = center_alignment
        ws.cell(row=row_idx, column=7).alignment = center_alignment
    
    # Auto-adjust column widths
    for col in range(1, 8):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add legend at the bottom
    last_row = start_row + len(students_with_percentage) + 3
    ws.merge_cells(f'A{last_row}:G{last_row}')
    ws[f'A{last_row}'] = "LEGEND: Good Attendance (≥75%) | Attendance Shortage (<75%)"
    ws[f'A{last_row}'].font = Font(italic=True)
    ws[f'A{last_row}'].alignment = center_alignment
    
    # Create response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=attendance_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response

@app.route('/admin/delete-student', methods=['POST'])
@login_required
@admin_required
def delete_student():
    student_id = request.form.get('student_id')
    
    print(f"DELETE DEBUG: student_id = {student_id}")
    print(f"DELETE DEBUG: form data = {dict(request.form)}")
    
    if not student_id:
        flash('No student ID provided', 'danger')
        return redirect(url_for('view_students'))
    
    try:
        conn = get_db_connection()
        
        # Check if student exists
        student = conn.execute('SELECT * FROM students WHERE id = ?', (student_id,)).fetchone()
        if not student:
            flash(f'Student with ID {student_id} not found', 'danger')
            conn.close()
            return redirect(url_for('view_students'))
        
        print(f"DELETE DEBUG: Found student {student['name']}")
        
        # Delete attendance records first
        conn.execute('DELETE FROM attendance WHERE student_id = ?', (student_id,))
        print("DELETE DEBUG: Deleted attendance records")
        
        # Delete student
        conn.execute('DELETE FROM students WHERE id = ?', (student_id,))
        print("DELETE DEBUG: Deleted student record")
        
        # Delete user account
        conn.execute('DELETE FROM users WHERE id = ?', (student['user_id'],))
        print("DELETE DEBUG: Deleted user record")
        
        # Delete image file
        if student['image_path']:
            image_path = os.path.join(app.config['UPLOAD_FOLDER'], student['image_path'])
            if os.path.exists(image_path):
                os.remove(image_path)
                print(f"DELETE DEBUG: Deleted image {image_path}")
        
        conn.commit()
        conn.close()
        
        flash(f'Student {student["name"]} deleted successfully!', 'success')
        print(f"DELETE DEBUG: Successfully deleted {student['name']}")
        
    except Exception as e:
        flash(f'Error deleting student: {str(e)}', 'danger')
        print(f"DELETE DEBUG: Error - {str(e)}")
    
    return redirect(url_for('view_students'))



# Add these routes to your app.py file

# Admin route to view attendance percentages of all students
@app.route('/admin/attendance-report')
@login_required
@admin_required
def attendance_report():
    conn = get_db_connection()
    
    # Get all students with their attendance data
    students_data = conn.execute('''
        SELECT s.id, s.name, s.roll_number, s.email,
               COUNT(a.id) as days_present
        FROM students s
        LEFT JOIN attendance a ON s.id = a.student_id
        GROUP BY s.id, s.name, s.roll_number, s.email
        ORDER BY s.roll_number
    ''').fetchall()
    
    # Calculate total working days (you can modify this logic based on your needs)
    # For now, let's count unique dates in attendance table as total working days
    total_days_result = conn.execute('SELECT COUNT(DISTINCT date) as total_days FROM attendance').fetchone()
    total_working_days = max(total_days_result['total_days'], 1)  # Avoid division by zero
    
    # If no attendance records exist, use a default value or calculate based on semester
    if total_working_days == 0:
        # You can set this to your actual working days per semester
        total_working_days = 90  # Example: 90 working days in a semester
    
    conn.close()
    
    # Calculate percentages
    students_with_percentage = []
    for student in students_data:
        percentage = (student['days_present'] / total_working_days) * 100
        students_with_percentage.append({
            'id': student['id'],
            'name': student['name'],
            'roll_number': student['roll_number'],
            'email': student['email'],
            'days_present': student['days_present'],
            'total_days': total_working_days,
            'percentage': round(percentage, 2),
            'is_shortage': percentage < 75
        })
    
    return render_template('admin/attendance_report.html', 
                          students=students_with_percentage, 
                          total_working_days=total_working_days)


@app.route('/about')
def about():
    return render_template('about.html')
@app.route('/contactus')
def contactus():
    return render_template('contactus.html')





# Student route to view their own attendance percentage
@app.route('/student/percentage')
@login_required
def student_attendance_percentage():
    if session['role'] != 'student':
        return redirect(url_for('admin_dashboard'))
    
    conn = get_db_connection()
    
    # Get student information
    student = conn.execute('''
        SELECT * FROM students WHERE user_id = ?
    ''', (session['user_id'],)).fetchone()
    
    # Get student's attendance count
    attendance_data = conn.execute('''
        SELECT COUNT(*) as days_present
        FROM attendance 
        WHERE student_id = ?
    ''', (student['id'],)).fetchone()
    
    # Get total working days
    total_days_result = conn.execute('SELECT COUNT(DISTINCT date) as total_days FROM attendance').fetchone()
    total_working_days = max(total_days_result['total_days'], 1)
    
    # If no attendance records exist, use a default value
    if total_working_days == 0:
        total_working_days = 90  # Default working days
    
    # Get detailed attendance records for the calendar view
    attendance_records = conn.execute('''
        SELECT date, time FROM attendance
        WHERE student_id = ?
        ORDER BY date DESC
    ''', (student['id'],)).fetchall()
    
    conn.close()
    
    # Calculate percentage
    days_present = attendance_data['days_present']
    percentage = (days_present / total_working_days) * 100
    is_shortage = percentage < 75
    
    # Calculate days needed to reach 75%
    days_needed = 0
    if is_shortage:
        days_needed = max(0, int((0.75 * total_working_days) - days_present))
    
    return render_template('student/percentage.html',
                          student=student,
                          days_present=days_present,
                          total_working_days=total_working_days,
                          percentage=round(percentage, 2),
                          is_shortage=is_shortage,
                          days_needed=days_needed,
                          attendance_records=attendance_records)

# API route to get attendance statistics (optional - for AJAX updates)
@app.route('/api/attendance-stats')
@login_required
def attendance_stats():
    conn = get_db_connection()
    
    if session['role'] == 'admin':
        # Admin gets overall statistics
        stats = conn.execute('''
            SELECT 
                COUNT(DISTINCT s.id) as total_students,
                COUNT(DISTINCT a.date) as total_working_days,
                COUNT(a.id) as total_attendance_records,
                AVG(attendance_count.days_present) as avg_attendance
            FROM students s
            LEFT JOIN attendance a ON s.id = a.student_id
            LEFT JOIN (
                SELECT student_id, COUNT(*) as days_present
                FROM attendance 
                GROUP BY student_id
            ) attendance_count ON s.id = attendance_count.student_id
        ''').fetchone()
        
        conn.close()
        return jsonify({
            'total_students': stats['total_students'],
            'total_working_days': stats['total_working_days'],
            'total_attendance_records': stats['total_attendance_records'],
            'average_attendance': round(stats['avg_attendance'] or 0, 2)
        })
    
    else:
        # Student gets their own statistics
        student = conn.execute('SELECT * FROM students WHERE user_id = ?', 
                              (session['user_id'],)).fetchone()
        
        student_stats = conn.execute('''
            SELECT 
                COUNT(*) as days_present,
                (SELECT COUNT(DISTINCT date) FROM attendance) as total_working_days
            FROM attendance 
            WHERE student_id = ?
        ''', (student['id'],)).fetchone()
        
        conn.close()
        
        total_days = max(student_stats['total_working_days'], 1)
        percentage = (student_stats['days_present'] / total_days) * 100
        
        return jsonify({
            'days_present': student_stats['days_present'],
            'total_working_days': total_days,
            'percentage': round(percentage, 2),
            'is_shortage': percentage < 75
        })

# API route to get the current WiFi information
@app.route('/api/wifi-info')
@login_required
def wifi_info():
    from utils.wifi_checker import get_current_wifi_info
    info = get_current_wifi_info()
    return jsonify(info)

# Route for setting the WiFi access point
@app.route('/api/set-access-point', methods=['POST'])
@login_required
@admin_required
def set_access_point():
    data = request.json
    ssid = data.get('ssid')
    bssid = data.get('bssid')
    description = data.get('description', '')
    
    if not ssid or not bssid:
        return jsonify({'success': False, 'message': 'SSID and BSSID are required'})
    
    try:
        conn = get_db_connection()
        
        # First delete any existing access points (we only want one required AP)
        conn.execute('DELETE FROM allowed_access_points')
        
        # Insert the new access point
        conn.execute(
            'INSERT INTO allowed_access_points (ssid, bssid, description) VALUES (?, ?, ?)',
            (ssid, bssid, description)
        )
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Access point set successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out', 'success')
    return redirect(url_for('index'))

# Admin routes
@app.route('/admin/dashboard')
@login_required
@admin_required
def admin_dashboard():
    conn = get_db_connection()
    student_count = conn.execute('SELECT COUNT(*) as count FROM students').fetchone()['count']
    today = datetime.now().strftime('%Y-%m-%d')
    attendance_today = conn.execute('SELECT COUNT(*) as count FROM attendance WHERE date = ?', (today,)).fetchone()['count']
    conn.close()
    
    return render_template('admin/dashboard.html', student_count=student_count, attendance_today=attendance_today)

@app.route('/admin/students')
@login_required
@admin_required
def view_students():
    conn = get_db_connection()
    students = conn.execute('SELECT * FROM students').fetchall()
    conn.close()
    
    return render_template('admin/students.html', students=students)

@app.route('/admin/register', methods=['GET', 'POST'])
@login_required
@admin_required
def register_student():
    if request.method == 'POST':
        name = request.form['name']
        roll_number = request.form['roll_number']
        email = request.form['email']
        
        # Generate a random password
        password = str(uuid.uuid4().hex)[:8]
        hashed_password = generate_password_hash(password)
        
        # Check if student exists
        conn = get_db_connection()
        existing_student = conn.execute('SELECT * FROM students WHERE email = ? OR roll_number = ?', 
                                       (email, roll_number)).fetchone()
        
        if existing_student:
            conn.close()
            flash('Student with this email or roll number already exists', 'danger')
            return redirect(url_for('register_student'))
        
        # Handle the image upload
        if 'image' not in request.files:
            conn.close()
            flash('No image part', 'danger')
            return redirect(request.url)
            
        file = request.files['image']
        
        if file.filename == '':
            conn.close()
            flash('No selected file', 'danger')
            return redirect(request.url)
            
        if file and allowed_file(file.filename):
            filename = secure_filename(f"{roll_number}_{uuid.uuid4().hex}.jpg")
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            # Create user account
            conn.execute(
                'INSERT INTO users (email, password, role) VALUES (?, ?, ?)',
                (email, hashed_password, 'student')
            )
            user_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            
            # Create student profile
            conn.execute(
                'INSERT INTO students (user_id, name, roll_number, email, image_path) VALUES (?, ?, ?, ?, ?)',
                (user_id, name, roll_number, email, filename)
            )
            conn.commit()
            conn.close()
            
            flash(f'Student registered successfully! Temporary password: {password}', 'success')
            return redirect(url_for('view_students'))
    
    return render_template('admin/register.html')

# Updated WiFi settings route
@app.route('/admin/settings', methods=['GET', 'POST'])
@login_required
@admin_required
def wifi_settings():
    conn = get_db_connection()
    
    # Check if the allowed_access_points table exists
    cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='allowed_access_points'")
    table_exists = cursor.fetchone() is not None
    
    # If table doesn't exist, create it
    if not table_exists:
        conn.executescript('''
            -- Add new table to store allowed access points
            CREATE TABLE IF NOT EXISTS allowed_access_points (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ssid TEXT NOT NULL,
                bssid TEXT NOT NULL,  -- MAC address of the access point
                description TEXT,     -- Optional description (e.g., "Library AP", "Classroom Building")
                UNIQUE(bssid)
            );
        ''')
        conn.commit()
    
    settings = conn.execute('SELECT * FROM settings').fetchone()
    
    # Handle form submissions
    if request.method == 'POST':
        if 'update_settings' in request.form:
            # Update general WiFi settings
            wifi_ssid = request.form['wifi_ssid']
            wifi_password = request.form['wifi_password']
            
            conn.execute(
                'UPDATE settings SET wifi_ssid = ?, wifi_password = ?',
                (wifi_ssid, wifi_password)
            )
            conn.commit()
            flash('WiFi settings updated successfully', 'success')
            
        elif 'delete_ap' in request.form:
            # Delete access point
            ap_id = request.form['ap_id']
            conn.execute('DELETE FROM allowed_access_points WHERE id = ?', (ap_id,))
            conn.commit()
            flash('Access point requirement removed', 'success')
    
    # Get the required access point (if set)
    required_ap = conn.execute('SELECT * FROM allowed_access_points LIMIT 1').fetchone()
    
    conn.close()
    
    return render_template('admin/settings.html', settings=settings, required_ap=required_ap)

# API route to check WiFi connection for attendance
@app.route('/api/check-wifi', methods=['POST'])
@login_required
def check_wifi():
    try:
        if 'role' not in session or session['role'] != 'student':
            return jsonify({'success': False, 'message': 'Unauthorized'})
        
        conn = get_db_connection()
        
        # Use the simplified verification
        from utils.wifi_checker import check_wifi_connection
        wifi_result = check_wifi_connection(conn)
        
        conn.close()
        
        return jsonify({
            'success': True,
            'connected': wifi_result.get('connected', False),
            'ssid': wifi_result.get('ssid'),
            'bssid': wifi_result.get('bssid'),
            'required_bssid': wifi_result.get('required_bssid'),
            'message': wifi_result.get('message')
        })
    except Exception as e:
        app.logger.error(f"WiFi check error: {str(e)}")
        return jsonify({'success': False, 'message': f"Server error: {str(e)}"})

# New endpoint to get MAC address
@app.route('/api/mac-address', methods=['GET'])
def mac_address():
    mac = get_wifi_mac()
    return jsonify({'mac_address': mac})

# FIXED ROUTE - Combined the duplicate route into one function
@app.route('/api/mark-attendance', methods=['POST'])
@login_required
def process_attendance():
    if session['role'] != 'student':
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    # Get the image data from the request
    if 'image' not in request.files:
        return jsonify({'success': False, 'message': 'No image provided'})
    
    image_file = request.files['image']
    mac_address = request.form.get('mac_address', '')
    
    # Read the image
    in_memory_file = image_file.read()
    nparr = np.frombuffer(in_memory_file, np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    
    conn = get_db_connection()
    student = conn.execute('SELECT * FROM students WHERE user_id = ?', 
                          (session['user_id'],)).fetchone()
    
    # Check if attendance already marked today
    today = datetime.now().strftime('%Y-%m-%d')
    already_marked = conn.execute('''
        SELECT COUNT(*) as count FROM attendance 
        WHERE student_id = ? AND date = ?
    ''', (student['id'], today)).fetchone()['count'] > 0
    
    if already_marked:
        conn.close()
        return jsonify({'success': False, 'message': 'Attendance already marked today'})
    
    # Get the required access point (if set)
    required_ap = conn.execute('SELECT * FROM allowed_access_points LIMIT 1').fetchone()
    
    # Verify WiFi connection
    try:
        from utils.wifi_checker import check_wifi_connection
        wifi_result = check_wifi_connection(conn)
        
        # Check if MAC address matches directly when WiFi detection fails
        mac_match = False
        if required_ap and mac_address:
            # Normalize both MAC addresses for comparison
            from utils.wifi_checker import normalize_mac
            normalized_mac = normalize_mac(mac_address)
            normalized_required = normalize_mac(required_ap['bssid'])
            
            if normalized_mac and normalized_required and normalized_mac == normalized_required:
                mac_match = True
                app.logger.info(f"MAC address match: {normalized_mac} = {normalized_required}")
        
        # If there's a required AP and WiFi is not connected and MAC doesn't match
        if required_ap and not wifi_result['connected'] and not mac_match:
            error_message = f"You must be connected to the required WiFi access point to mark attendance. Please connect to {required_ap['ssid']} network."
            conn.close()
            return jsonify({
                'success': False, 
                'message': error_message
            })
    except Exception as e:
        app.logger.error(f"Error verifying WiFi: {str(e)}")
        # If there's an MAC address match, we can proceed even if WiFi check failed
        if not mac_match and required_ap:
            conn.close()
            return jsonify({
                'success': False,
                'message': 'Could not verify WiFi connection. Please try again.'
            })
    
    # Get the reference image path
    reference_image_path = os.path.join(app.config['UPLOAD_FOLDER'], student['image_path'])
    
    # Compare faces
    try:
        match = compare_faces(img, reference_image_path)
    except Exception as e:
        app.logger.error(f"Error comparing faces: {str(e)}")
        conn.close()
        return jsonify({
            'success': False,
            'message': f'Face verification error: {str(e)}'
        })
    
    if match:
        # Mark attendance
        now = datetime.now()
        date = now.strftime('%Y-%m-%d')
        time = now.strftime('%H:%M:%S')
        
        # Insert attendance record
        conn.execute('''
            INSERT INTO attendance (student_id, date, time)
            VALUES (?, ?, ?)
        ''', (student['id'], date, time))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Attendance marked successfully!'
        })
    else:
        conn.close()
        return jsonify({
            'success': False,
            'message': 'Face verification failed. Please ensure good lighting and position your face clearly in front of the camera.'
        })

@app.route('/admin/attendance')
@login_required
@admin_required
def view_attendance():
    date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    
    conn = get_db_connection()
    attendance_records = conn.execute('''
        SELECT a.id, s.name, s.roll_number, a.date, a.time
        FROM attendance a
        JOIN students s ON a.student_id = s.id
        WHERE a.date = ?
        ORDER BY a.time DESC
    ''', (date,)).fetchall()
    conn.close()
    
    return render_template('admin/attendance.html', attendance_records=attendance_records, selected_date=date)

# Student routes
@app.route('/student/dashboard')
@login_required
def student_dashboard():
    if session['role'] != 'student':
        return redirect(url_for('admin_dashboard'))
    
    conn = get_db_connection()
    student = conn.execute('''
        SELECT * FROM students WHERE user_id = ?
    ''', (session['user_id'],)).fetchone()
    
    # Get attendance count
    attendance_count = conn.execute('''
        SELECT COUNT(*) as count FROM attendance WHERE student_id = ?
    ''', (student['id'],)).fetchone()['count']
    
    # Check if attendance marked today
    today = datetime.now().strftime('%Y-%m-%d')
    marked_today = conn.execute('''
        SELECT COUNT(*) as count FROM attendance 
        WHERE student_id = ? AND date = ?
    ''', (student['id'], today)).fetchone()['count'] > 0
    
    conn.close()
    
    return render_template('student/dashboard.html', 
                          student=student, 
                          attendance_count=attendance_count,
                          marked_today=marked_today)

@app.route('/student/attendance')
@login_required
def mark_attendance():
    if session['role'] != 'student':
        return redirect(url_for('admin_dashboard'))
    
    conn = get_db_connection()
    student = conn.execute('SELECT * FROM students WHERE user_id = ?', 
                          (session['user_id'],)).fetchone()
    
    # Check if attendance already marked today
    today = datetime.now().strftime('%Y-%m-%d')
    attendance_record = conn.execute('''
        SELECT * FROM attendance 
        WHERE student_id = ? AND date = ?
        ORDER BY time DESC
        LIMIT 1
    ''', (student['id'], today)).fetchone()
    
    already_marked = attendance_record is not None
    last_marked_time = attendance_record['time'] if already_marked else None
    
    # Get WiFi settings
    settings = conn.execute('SELECT * FROM settings').fetchone()
    conn.close()
    
    return render_template('student/attendance.html', 
                          student=student,
                          already_marked=already_marked,
                          last_marked_time=last_marked_time,
                          wifi_ssid=settings['wifi_ssid'])

@app.route('/student/history')
@login_required
def attendance_history():
    if session['role'] != 'student':
        return redirect(url_for('admin_dashboard'))
    
    conn = get_db_connection()
    student = conn.execute('SELECT * FROM students WHERE user_id = ?', 
                          (session['user_id'],)).fetchone()
    
    attendance_records = conn.execute('''
        SELECT * FROM attendance
        WHERE student_id = ?
        ORDER BY date DESC, time DESC
    ''', (student['id'],)).fetchall()
    
    conn.close()
    
    return render_template('student/history.html', 
                          student=student,
                          attendance_records=attendance_records)

if __name__ == '__main__':
    # Create database tables if they don't exist
    if not os.path.exists('database'):
        os.makedirs('database')
    
    if not os.path.exists('database/attendance.db'):
        init_db()
        
    app.run(debug=True)